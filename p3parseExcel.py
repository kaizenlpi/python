''' 
Author: Matt Brady
Written in python 3.6.8
Purpose: This script will parse row 301 to row 6197 in column 2 to get the computer hostnames.
It will not print out the unicode formatting like u'(somevalue)

'''

#!/bin/user/env python3
import openpyxl
wb = openpyxl.load_workbook('AnsibleReadinessReviewIP.xlsx')
sheet = wb['CIsIPaddresses']
sheet['A301']
sheet['A301'].value
A = sheet['A301']
A.value
sheet.cell(row=301, column=1)
# Get the row, column, and vlaue from the cell
'Row %s, Column %s is %s' % (A.row, A.column, A.value)
'Cell %s is %s' % (A.coordinate, A.value)
sheet['A301']
for i in range(301, 6197):
    print(i,sheet.cell(row=i, column=2).value)

